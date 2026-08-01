import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Setup styles
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy Blue
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

section_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Soft Blue
section_font = Font(name="Calibri", size=11, bold=True, color="1E3A8A")

title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
bold_font = Font(name="Calibri", size=10, bold=True)
regular_font = Font(name="Calibri", size=10)

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ----------------------------------------------------
# Sheet 1: Implementation Plan
# ----------------------------------------------------
ws1 = wb.active
ws1.title = "Implementation Plan"
ws1.views.sheetView[0].showGridLines = True

ws1.cell(row=1, column=1, value="EduFeedback System - Implementation Plan & Roadmap").font = title_font
ws1.row_dimensions[1].height = 30

headers1 = ["Phase", "Phase Name", "Target Modules & Components", "Key Tasks & Technical Steps", "Status", "Expected Output"]
ws1.row_dimensions[3].height = 25

for col_num, h in enumerate(headers1, 1):
    cell = ws1.cell(row=3, column=col_num, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center

plan_data = [
    ("Phase 1", "UI & Frontend Foundation", "Landing Page, Student Feedback Wizard, Admin Login, Responsive Navigation", 
     "Setup Vite + React project, Tailwind CSS, Lucide icons, setup App router, craft responsive UI components, step wizard layout.", "Planned", "Complete, styled frontend skeleton and pages"),
    
    ("Phase 2", "Database & Schema Setup", "Supabase PostgreSQL Schema, RLS Policies, Seed Data, Transaction Functions", 
     "Write supabase/schema.sql script with 11 tables, foreign keys, indexes, submit_feedback_atomic RPC function, RLS security policies.", "Planned", "Production SQL setup script & local data service"),
    
    ("Phase 3", "Student Feedback Module", "Multi-step Feedback Portal, Dependent Dropdowns, Rating Scale, Review Step", 
     "Implement 4-step wizard: Academic Selection -> 15 Rating Questions -> Student Suggestions -> Review & Atomic DB Submission.", "Planned", "Fully functional student feedback submission portal"),
    
    ("Phase 4", "Master Data Management", "CRUD for Academic Years, Depts, Programmes, Faculty, Courses, Mappings", 
     "Build admin management interfaces with search, filtering, modal forms, activation toggles, and soft deletion protection.", "Planned", "Complete admin master data CRUD suite"),
    
    ("Phase 5", "Analytics & Dashboard", "KPI Metrics Cards, 5 Recharts Visualizations, Global Filtering Toolbar", 
     "Build KPI calculations, Faculty averages, Department bar chart, Grade donut chart, Q1-Q15 breakdown, trend line chart, global filter state.", "Planned", "Real-time interactive administrative analytics dashboard"),
    
    ("Phase 6", "IQAC Reports & Export", "Interactive Report Generator, Printable PDF View (@media print), CSV Export", 
     "Implement filtered report view, printable CSS layout formatted for institutional headers, signature blocks, browser print & CSV download.", "Planned", "IQAC compliant report generation & export engine"),
    
    ("Phase 7", "Security & Quality Assurance", "Auth Protection, RLS Enforcement, Form Validation, Mobile Responsive QA", 
     "Protect admin routes, enforce validation rules, test duplicate submission handling, verify mobile touch targets & error boundaries.", "Planned", "Tested, secure, production-ready web application")
]

for row_idx, data in enumerate(plan_data, start=4):
    ws1.row_dimensions[row_idx].height = 40
    for col_idx, value in enumerate(data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = bold_font if col_idx in [1, 2] else regular_font
        cell.alignment = align_center if col_idx in [1, 5] else align_left
        cell.border = thin_border
        if col_idx == 5:
            cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid") # Yellow highlight for Planned

# ----------------------------------------------------
# Sheet 2: Database Schema
# ----------------------------------------------------
ws2 = wb.create_sheet(title="Database Schema")
ws2.views.sheetView[0].showGridLines = True

ws2.cell(row=1, column=1, value="EduFeedback System - Database Schema & Data Model").font = title_font
ws2.row_dimensions[1].height = 30

headers2 = ["Table Name", "Column Name", "Data Type", "Constraints & Foreign Keys", "Description"]
ws2.row_dimensions[3].height = 25

for col_num, h in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col_num, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center

schema_data = [
    ("profiles", "id", "UUID", "PRIMARY KEY", "Unique profile identifier"),
    ("profiles", "user_id", "UUID", "FK -> auth.users(id)", "Supabase Auth user link"),
    ("profiles", "full_name", "TEXT", "NOT NULL", "Admin full name"),
    ("profiles", "role", "TEXT", "DEFAULT 'admin'", "User access role"),
    
    ("academic_years", "id", "UUID", "PRIMARY KEY", "Academic year ID"),
    ("academic_years", "year_name", "VARCHAR(20)", "UNIQUE, NOT NULL", "Year label (e.g. 2026-2027)"),
    ("academic_years", "is_current", "BOOLEAN", "DEFAULT false", "Current active academic year flag"),
    ("academic_years", "status", "VARCHAR(20)", "DEFAULT 'active'", "Active/inactive status"),
    
    ("departments", "id", "UUID", "PRIMARY KEY", "Department ID"),
    ("departments", "department_code", "VARCHAR(20)", "UNIQUE, NOT NULL", "Dept code (e.g. CSE)"),
    ("departments", "department_name", "TEXT", "NOT NULL", "Full department name"),
    ("departments", "status", "VARCHAR(20)", "DEFAULT 'active'", "Active/inactive status"),
    
    ("programmes", "id", "UUID", "PRIMARY KEY", "Programme ID"),
    ("programmes", "programme_code", "VARCHAR(20)", "UNIQUE, NOT NULL", "Programme code (e.g. BE-CSE)"),
    ("programmes", "programme_name", "TEXT", "NOT NULL", "Full degree programme name"),
    ("programmes", "department_id", "UUID", "FK -> departments(id)", "Parent department link"),
    ("programmes", "semesters_count", "INTEGER", "DEFAULT 8", "Total number of semesters"),
    
    ("faculty", "id", "UUID", "PRIMARY KEY", "Faculty ID"),
    ("faculty", "faculty_code", "VARCHAR(20)", "UNIQUE, NOT NULL", "Faculty employee ID"),
    ("faculty", "faculty_name", "TEXT", "NOT NULL", "Faculty full name"),
    ("faculty", "email", "TEXT", "UNIQUE", "Faculty official email"),
    ("faculty", "department_id", "UUID", "FK -> departments(id)", "Department link"),
    ("faculty", "designation", "TEXT", "NOT NULL", "Academic title/designation"),
    
    ("courses", "id", "UUID", "PRIMARY KEY", "Course ID"),
    ("courses", "course_code", "VARCHAR(20)", "UNIQUE, NOT NULL", "Course code (e.g. CS301)"),
    ("courses", "course_title", "TEXT", "NOT NULL", "Course title"),
    ("courses", "department_id", "UUID", "FK -> departments(id)", "Department link"),
    ("courses", "programme_id", "UUID", "FK -> programmes(id)", "Programme link"),
    ("courses", "semester", "INTEGER", "NOT NULL", "Semester number"),
    
    ("faculty_course_mappings", "id", "UUID", "PRIMARY KEY", "Mapping ID"),
    ("faculty_course_mappings", "academic_year_id", "UUID", "FK -> academic_years(id)", "Academic year link"),
    ("faculty_course_mappings", "faculty_id", "UUID", "FK -> faculty(id)", "Assigned faculty member"),
    ("faculty_course_mappings", "course_id", "UUID", "FK -> courses(id)", "Assigned course"),
    ("faculty_course_mappings", "section", "VARCHAR(10)", "DEFAULT 'A'", "Class section (A, B, C)"),
    
    ("feedback", "id", "UUID", "PRIMARY KEY", "Feedback submission ID"),
    ("feedback", "academic_year_id", "UUID", "FK -> academic_years(id)", "Academic year reference"),
    ("feedback", "department_id", "UUID", "FK -> departments(id)", "Department reference"),
    ("feedback", "programme_id", "UUID", "FK -> programmes(id)", "Programme reference"),
    ("feedback", "semester", "INTEGER", "NOT NULL", "Semester number"),
    ("feedback", "course_id", "UUID", "FK -> courses(id)", "Course reference"),
    ("feedback", "faculty_id", "UUID", "FK -> faculty(id)", "Faculty evaluated"),
    ("feedback", "mapping_id", "UUID", "FK -> faculty_course_mappings(id)", "Mapping record reference"),
    ("feedback", "suggestion_positive", "TEXT", "NULLABLE", "What student liked most"),
    ("feedback", "suggestion_improvement", "TEXT", "NULLABLE", "Suggested improvements"),
    ("feedback", "additional_comments", "TEXT", "NULLABLE", "Additional comments"),
    ("feedback", "submitted_at", "TIMESTAMPTZ", "DEFAULT NOW()", "Submission timestamp"),
    
    ("feedback_ratings", "id", "UUID", "PRIMARY KEY", "Rating item ID"),
    ("feedback_ratings", "feedback_id", "UUID", "FK -> feedback(id) ON DELETE CASCADE", "Parent feedback reference"),
    ("feedback_ratings", "question_number", "INTEGER", "CHECK (1-15)", "Question criteria number"),
    ("feedback_ratings", "rating", "INTEGER", "CHECK (1-5)", "Score rating assigned")
]

for row_idx, data in enumerate(schema_data, start=4):
    ws2.row_dimensions[row_idx].height = 22
    for col_idx, value in enumerate(data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = align_center if col_idx in [1, 2, 3] else align_left
        cell.border = thin_border
        cell.font = bold_font if col_idx in [1, 2] else regular_font

# ----------------------------------------------------
# Sheet 3: Evaluation Questions
# ----------------------------------------------------
ws3 = wb.create_sheet(title="Evaluation Questions")
ws3.views.sheetView[0].showGridLines = True

ws3.cell(row=1, column=1, value="EduFeedback System - 15 Evaluation Criteria").font = title_font
ws3.row_dimensions[1].height = 30

headers3 = ["Q.No", "Evaluation Criteria", "Category", "Rating Scale", "Max Score"]
ws3.row_dimensions[3].height = 25

for col_num, h in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col_num, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center

questions_data = [
    ("Q1", "Faculty comes to class regularly and punctually.", "Punctuality & Discipline", "1 (Poor) to 5 (Excellent)", 5),
    ("Q2", "Faculty demonstrates thorough subject knowledge and command over the course.", "Subject Command", "1 (Poor) to 5 (Excellent)", 5),
    ("Q3", "Concepts are explained clearly and systematically.", "Clarity & Pedagogy", "1 (Poor) to 5 (Excellent)", 5),
    ("Q4", "Course outcomes and objectives are explained clearly.", "Curriculum Clarity", "1 (Poor) to 5 (Excellent)", 5),
    ("Q5", "Teaching methodology is effective.", "Teaching Effectiveness", "1 (Poor) to 5 (Excellent)", 5),
    ("Q6", "Faculty effectively uses ICT tools such as PPT, videos, smart boards, simulations and LMS.", "ICT & Digital Tools", "1 (Poor) to 5 (Excellent)", 5),
    ("Q7", "Faculty encourages student participation and interaction.", "Classroom Engagement", "1 (Poor) to 5 (Excellent)", 5),
    ("Q8", "Faculty explains concepts using real-life and industry examples.", "Practical Application", "1 (Poor) to 5 (Excellent)", 5),
    ("Q9", "Faculty completes the syllabus according to the academic schedule.", "Syllabus Completion", "1 (Poor) to 5 (Excellent)", 5),
    ("Q10", "Faculty conducts quizzes, assignments and tutorials regularly.", "Continuous Assessment", "1 (Poor) to 5 (Excellent)", 5),
    ("Q11", "Faculty provides timely and constructive feedback on assignments and tests.", "Constructive Feedback", "1 (Poor) to 5 (Excellent)", 5),
    ("Q12", "Internal assessment is fair and transparent.", "Assessment Fairness", "1 (Poor) to 5 (Excellent)", 5),
    ("Q13", "Faculty is available for academic guidance outside regular class hours.", "Accessibility & Mentoring", "1 (Poor) to 5 (Excellent)", 5),
    ("Q14", "Faculty motivates students toward higher-order thinking and problem solving.", "Student Motivation", "1 (Poor) to 5 (Excellent)", 5),
    ("Q15", "Overall effectiveness of teaching.", "Overall Evaluation", "1 (Poor) to 5 (Excellent)", 5),
]

for row_idx, data in enumerate(questions_data, start=4):
    ws3.row_dimensions[row_idx].height = 25
    for col_idx, value in enumerate(data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = align_center if col_idx in [1, 3, 4, 5] else align_left
        cell.border = thin_border
        cell.font = bold_font if col_idx == 1 else regular_font

# ----------------------------------------------------
# Sheet 4: Grading & IQAC Standards
# ----------------------------------------------------
ws4 = wb.create_sheet(title="Grading & IQAC Standards")
ws4.views.sheetView[0].showGridLines = True

ws4.cell(row=1, column=1, value="EduFeedback System - Grading & IQAC Evaluation Standards").font = title_font
ws4.row_dimensions[1].height = 30

ws4.cell(row=3, column=1, value="1. Institutional Grade Scale (Total Score out of 75)").font = section_font
headers4_1 = ["Score Range", "Grade", "Performance Classification", "Action Level"]
ws4.row_dimensions[4].height = 25

for col_num, h in enumerate(headers4_1, 1):
    cell = ws4.cell(row=4, column=col_num, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center

grades_data = [
    ("68 – 75", "A+", "Outstanding", "Commendation & Exemplary Teaching"),
    ("60 – 67", "A", "Excellent", "Highly Satisfactory"),
    ("52 – 59", "B+", "Very Good", "Meets High Standards"),
    ("45 – 51", "B", "Good", "Satisfactory Performance"),
    ("37 – 44", "C", "Satisfactory", "Minor Guidance Recommended"),
    ("Below 37", "D", "Needs Improvement", "Mandatory Teaching Review & Action Plan")
]

for row_idx, data in enumerate(grades_data, start=5):
    ws4.row_dimensions[row_idx].height = 22
    for col_idx, value in enumerate(data, start=1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = align_center if col_idx in [1, 2] else align_left
        cell.border = thin_border
        cell.font = bold_font if col_idx in [1, 2] else regular_font

ws4.cell(row=12, column=1, value="2. IQAC Percentage-Based Performance Interpretation").font = section_font
headers4_2 = ["Percentage Range", "IQAC Performance Level", "Status Indicator"]
ws4.row_dimensions[13].height = 25

for col_num, h in enumerate(headers4_2, 1):
    cell = ws4.cell(row=13, column=col_num, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center

iqac_data = [
    ("90% – 100%", "Excellent Performance", "Positive / Green"),
    ("80% – 89.99%", "Very Good Performance", "Info / Blue"),
    ("70% – 79.99%", "Good Performance", "Info / Cyan"),
    ("60% – 69.99%", "Satisfactory Performance", "Warning / Yellow"),
    ("Below 60%", "Improvement Required", "Critical / Red")
]

for row_idx, data in enumerate(iqac_data, start=14):
    ws4.row_dimensions[row_idx].height = 22
    for col_idx, value in enumerate(data, start=1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = align_center if col_idx in [1, 3] else align_left
        cell.border = thin_border
        cell.font = bold_font if col_idx in [1, 2] else regular_font

# ----------------------------------------------------
# Sheet 5: Application Routes & Access
# ----------------------------------------------------
ws5 = wb.create_sheet(title="Routes & User Roles")
ws5.views.sheetView[0].showGridLines = True

ws5.cell(row=1, column=1, value="EduFeedback System - Routes & Authorization Matrix").font = title_font
ws5.row_dimensions[1].height = 30

headers5 = ["Route Path", "Page Component", "Access Level", "Key Capabilities & Views"]
ws5.row_dimensions[3].height = 25

for col_num, h in enumerate(headers5, 1):
    cell = ws5.cell(row=3, column=col_num, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = align_center

routes_data = [
    ("/", "LandingPage.tsx", "Public", "Hero section, institutional branding, quick links to Give Feedback or Admin Login"),
    ("/feedback", "FeedbackPage.tsx", "Public (Student)", "4-step feedback submission wizard with academic selection & 15 criteria"),
    ("/feedback/success", "FeedbackSuccessPage.tsx", "Public (Student)", "Submission confirmation message and option to submit another feedback"),
    ("/admin/login", "AdminLogin.tsx", "Public / Admin Login", "Secure authentication form with email, password, and demo quick-login button"),
    ("/admin/dashboard", "Dashboard.tsx", "Protected (Admin)", "6 KPI summary cards, 5 Recharts visualizations, global filters toolbar"),
    ("/admin/faculty-performance", "FacultyPerformance.tsx", "Protected (Admin)", "Searchable, sortable faculty metrics table with grade & performance badges"),
    ("/admin/faculty/:id", "FacultyDetail.tsx", "Protected (Admin)", "Detailed faculty score breakdown (Q1-Q15), strengths analysis, student comments"),
    ("/admin/department-performance", "DepartmentPerformance.tsx", "Protected (Admin)", "Department-level performance cards and comparative analytics"),
    ("/admin/responses", "FeedbackResponses.tsx", "Protected (Admin)", "List of submitted feedback records with filter, pagination & anonymous mode control"),
    ("/admin/reports", "ReportsPage.tsx", "Protected (Admin)", "IQAC printable analysis report view (@media print), PDF export, CSV download"),
    ("/admin/academic-years", "AcademicYears.tsx", "Protected (Admin)", "CRUD management for academic years & active year designation"),
    ("/admin/departments", "Departments.tsx", "Protected (Admin)", "CRUD management for academic departments"),
    ("/admin/programmes", "Programmes.tsx", "Protected (Admin)", "CRUD management for degree programmes & semester counts"),
    ("/admin/faculty", "FacultyPage.tsx", "Protected (Admin)", "CRUD management for faculty members"),
    ("/admin/courses", "CoursesPage.tsx", "Protected (Admin)", "CRUD management for academic courses"),
    ("/admin/mappings", "MappingsPage.tsx", "Protected (Admin)", "CRUD management for faculty-course-section assignments"),
    ("/admin/settings", "SettingsPage.tsx", "Protected (Admin)", "Institution configuration, feedback form open/close switch, anonymous mode toggle")
]

for row_idx, data in enumerate(routes_data, start=4):
    ws5.row_dimensions[row_idx].height = 22
    for col_idx, value in enumerate(data, start=1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = align_center if col_idx in [1, 3] else align_left
        cell.border = thin_border
        cell.font = bold_font if col_idx in [1, 2] else regular_font

# Adjust column widths automatically for all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 3] and sheet.title in ["Grading & IQAC Standards"] and cell.row == 12:
                continue
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_str = max(val_str.split('\n'), key=len)
            max_len = max(max_len, len(val_str))
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

output_path = r"c:\Users\HP\OneDrive\Desktop\feedback\dataform.xlsx"
wb.save(output_path)
print(f"Successfully generated {output_path}")
